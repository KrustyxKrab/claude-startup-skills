#!/usr/bin/env python3
"""
Shared Excel export utility for startup-research scripts.
Dependency: pip install openpyxl

Usage (from other scripts):
    from excel_utils import save_to_excel, check_openpyxl
    if check_openpyxl():
        save_to_excel("output.xlsx", {"Sheet1": rows, "Sheet2": rows})
"""

from __future__ import annotations


def check_openpyxl() -> bool:
    """Return True if openpyxl is available, print install hint otherwise."""
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        print("WARNING: openpyxl not installed — Excel export skipped.")
        print("  Install: pip install openpyxl --break-system-packages")
        return False


def save_to_excel(path: str, sheets: dict[str, list[dict]]) -> None:
    """
    Write a dict of {sheet_name: [row_dicts]} to an xlsx file.

    Each list of dicts becomes one sheet. The dict keys become the header row.
    Empty or None sheets are skipped.

    Args:
        path:   Output filename, e.g. "competitive-scan.xlsx"
        sheets: {"Competitors": [{"Name": "Acme", "Price": "$10"}], ...}
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL     = PatternFill("solid", fgColor="D9E1F2")   # light blue
    WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    for sheet_name, rows in sheets.items():
        if not rows:
            continue
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel max 31 chars

        headers = list(rows[0].keys())
        # Header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(vertical="center", horizontal="center")

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for col_idx, header in enumerate(headers, start=1):
                value = row.get(header, "")
                # Truncate very long strings for readability
                if isinstance(value, str) and len(value) > 500:
                    value = value[:497] + "..."
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = WRAP_ALIGN
                if fill:
                    cell.fill = fill

        # Auto-fit column widths (capped at 60)
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(header)),
                *[len(str(row.get(header, "") or "")[:80]) for row in rows],
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        ws.freeze_panes = "A2"  # freeze header row

    wb.save(path)
    print(f"Excel saved → {path}")
